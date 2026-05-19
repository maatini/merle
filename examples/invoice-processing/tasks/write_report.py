"""
Task: Write consolidated Excel report using pandas + openpyxl.

Demonstrates professional Excel output:
- Multiple sheets
- Formatting, formulas, conditional formatting
- Proper headers and types
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows

from merle_core import BaseTask


class WriteExcelReportTask(BaseTask):
    """Creates professional Excel output from enriched invoice data."""

    def __init__(self, settings: Any, output_dir: Path, invoices: list[dict[str, Any]]) -> None:
        super().__init__(settings, name="WriteExcelReport")
        self.output_dir = output_dir
        self.invoices = invoices
        self.output_dir.mkdir(parents=True, exist_ok=True)

    async def execute(self) -> dict[str, Any]:
        if not self.invoices:
            self.logger.warning("No invoices to generate report from")
            return {"report_path": None, "total_invoices": 0, "total_amount": 0.0}

        # Build DataFrame with explicit columns
        df = pd.DataFrame(self.invoices)
        columns_order = [
            "invoice_id", "supplier", "supplier_id", "invoice_date",
            "amount_gross", "amount_net", "vat_rate", "cost_center", "payment_terms"
        ]
        # Ensure all columns exist
        for col in columns_order:
            if col not in df.columns:
                df[col] = ""

        df = df[columns_order]

        # Map to user-friendly column headers
        df.columns = [
            "Invoice ID", "Supplier", "Supplier ID", "Invoice Date",
            "Amount Gross", "Amount Net", "VAT Rate", "Cost Center", "Payment Terms"
        ]

        # Create workbook
        wb = Workbook()

        # Detail sheet
        ws_detail = wb.active
        ws_detail.title = "Invoices"

        # Write DataFrame rows to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_detail.cell(row=r_idx, column=c_idx, value=value)

                # Header Styling
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

                cell.alignment = Alignment(horizontal="center", vertical="center")

        num_rows = len(df)
        # Write a total row in the detail sheet using native formulas
        tot_row = num_rows + 2
        ws_detail.cell(row=tot_row, column=4, value="Total:").font = Font(bold=True)

        cell_gross = ws_detail.cell(row=tot_row, column=5, value=f"=SUM(E2:E{num_rows+1})")
        cell_gross.font = Font(bold=True)
        cell_gross.number_format = '#,##0.00" €"'

        cell_net = ws_detail.cell(row=tot_row, column=6, value=f"=SUM(F2:F{num_rows+1})")
        cell_net.font = Font(bold=True)
        cell_net.number_format = '#,##0.00" €"'

        # Format data cells (numbers and percentages)
        for r in range(2, num_rows + 2):
            # Gross Amount
            cell_g = ws_detail.cell(row=r, column=5)
            cell_g.number_format = '#,##0.00" €"'
            cell_g.alignment = Alignment(horizontal="right")
            # Net Amount
            cell_n = ws_detail.cell(row=r, column=6)
            cell_n.number_format = '#,##0.00" €"'
            cell_n.alignment = Alignment(horizontal="right")
            # VAT Rate
            cell_v = ws_detail.cell(row=r, column=7)
            cell_v.number_format = '0.0%'
            cell_v.alignment = Alignment(horizontal="right")

        # Add Conditional Formatting to highlight high-value invoices (> 10,000 €)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        ws_detail.conditional_formatting.add(
            f"E2:E{num_rows+1}",
            CellIsRule(operator='greaterThan', formula=['10000'], fill=red_fill, font=red_font)
        )

        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.views.sheetView[0].showGridLines = True
        ws_summary["A1"] = "Invoice Processing Report"
        ws_summary["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws_summary["A3"] = "Generated"
        ws_summary["B3"] = pd.Timestamp.now().isoformat()

        # Use formulas pointing back to detail sheet
        summary = {
            "Total Invoices": f"=COUNTA(Invoices!A2:A{num_rows+1})",
            "Total Gross Amount (EUR)": f"=SUM(Invoices!E2:E{num_rows+1})",
            "Average Gross Amount (EUR)": f"=AVERAGE(Invoices!E2:E{num_rows+1})",
        }

        row = 5
        for key, value in summary.items():
            ws_summary[f"A{row}"] = key
            ws_summary[f"B{row}"] = value
            ws_summary[f"A{row}"].font = Font(bold=True)
            if "Amount" in key or "Total Gross" in key:
                ws_summary[f"B{row}"].number_format = '#,##0.00" €"'
            row += 1

        # Apply borders & auto-fit columns
        thin_border = Border(left=Side(style="thin", color="D3D3D3"),
                             right=Side(style="thin", color="D3D3D3"),
                             top=Side(style="thin", color="D3D3D3"),
                             bottom=Side(style="thin", color="D3D3D3"))

        for ws in [ws_detail, ws_summary]:
            for row_cells in ws.iter_rows():
                for cell in row_cells:
                    cell.border = thin_border
            # Auto-fit widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Write file
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M")
        report_path = self.output_dir / f"invoice_report_{timestamp}.xlsx"
        wb.save(report_path)

        self.logger.success("Excel report written to {}", report_path)
        return {
            "report_path": str(report_path),
            "total_invoices": len(df),
            "total_amount": float(df["Amount Gross"].sum()),
        }
