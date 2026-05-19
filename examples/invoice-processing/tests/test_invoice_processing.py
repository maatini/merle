"""
Tests for the Invoice Processing Bot tasks.
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pytest
from openpyxl import load_workbook

from tasks import (
    DownloadInvoicesTask,
    ParsePdfInvoicesTask,
    WriteExcelReportTask,
)
from scripts.generate_sample_pdfs import create_invoice_pdf

# Setup directories for testing
TEST_DIR = Path(__file__).resolve().parent / "test_data"


@pytest.fixture(autouse=True)
def setup_test_environment():
    """Create and clean up temporary test directories."""
    # Ensure test directories are clean
    if TEST_DIR.exists():
        shutil.rmtree(TEST_DIR)
    TEST_DIR.mkdir(parents=True, exist_ok=True)
    yield
    # Clean up after test run
    if TEST_DIR.exists():
        shutil.rmtree(TEST_DIR)


class MockSettings:
    def __init__(self):
        self.simulated_mode = True
        self.simulated_inbox_dir = TEST_DIR / "simulated_mail_inbox"
        self.input_dir = TEST_DIR / "invoices"
        self.output_dir = TEST_DIR / "reports"
        self.imap_host = ""
        self.imap_username = ""
        self.imap_password = ""


@pytest.mark.asyncio
async def test_pdf_parsing():
    """Test that ParsePdfInvoicesTask correctly parses reportlab-generated PDFs."""
    invoices_dir = TEST_DIR / "invoices"
    invoices_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = invoices_dir / "INV-2025-9999.pdf"

    # Create a test PDF with reportlab
    items = [("Cloud Infrastructure", 2, 500.00), ("Support Plan", 1, 100.00)]
    create_invoice_pdf(
        filename=pdf_path,
        invoice_id="INV-2025-9999",
        supplier="Test Supplier Ltd",
        date="2025-06-01",
        items=items,
        vat_rate=0.19,
    )

    task = ParsePdfInvoicesTask(MockSettings(), invoices_dir)
    res = await task.run()

    assert res["parsed"] == 1
    invoice = res["invoices"][0]
    assert invoice["invoice_id"] == "INV-2025-9999"
    assert invoice["supplier"] == "Test Supplier Ltd"
    assert invoice["amount_net"] == 1100.00
    assert invoice["amount_gross"] == 1100.00 * 1.19
    assert len(invoice["line_items"]) == 2
    assert invoice["line_items"][0]["description"] == "Cloud Infrastructure"
    assert invoice["line_items"][0]["qty"] == 2
    assert invoice["line_items"][0]["unit_price"] == 500.00
    assert invoice["line_items"][1]["description"] == "Support Plan"
    assert invoice["line_items"][1]["qty"] == 1
    assert invoice["line_items"][1]["unit_price"] == 100.00


@pytest.mark.asyncio
async def test_download_local_simulation():
    """Test DownloadInvoicesTask local email attachment extraction."""
    mock_settings = MockSettings()
    task = DownloadInvoicesTask(mock_settings, mock_settings.input_dir)

    # Calling run() should automatically populate simulated emails and process them
    res = await task.run()

    assert res["count"] == 3
    # Check that files were written to input_dir
    files = list(mock_settings.input_dir.glob("*.pdf"))
    assert len(files) == 3
    assert any(f.name == "INV-2025-0042.pdf" for f in files)

    # Check that processed emails were moved to archive
    archive_files = list((mock_settings.simulated_inbox_dir / "archive").glob("*.eml"))
    assert len(archive_files) == 3


@pytest.mark.asyncio
async def test_write_excel_report():
    """Test that WriteExcelReportTask correctly creates formatted sheets and formulas."""
    mock_settings = MockSettings()
    invoices = [
        {
            "invoice_id": "INV-2025-0001",
            "supplier": "Test Supplier",
            "supplier_id": "SUP-1234",
            "invoice_date": "2025-06-01",
            "amount_gross": 15000.00,
            "amount_net": 12605.04,
            "vat_rate": 0.19,
            "cost_center": "CC-100",
            "payment_terms": "30 days",
        }
    ]

    task = WriteExcelReportTask(mock_settings, mock_settings.output_dir, invoices)
    res = await task.run()

    report_path = Path(res["report_path"])
    assert report_path.exists()

    # Load workbook and check sheet names and basic structure
    wb = load_workbook(report_path)
    assert "Invoices" in wb.sheetnames
    assert "Summary" in wb.sheetnames

    # Check detail sheet headers and formulas
    ws_invoices = wb["Invoices"]
    assert ws_invoices["A1"].value == "Invoice ID"
    assert ws_invoices["E2"].value == 15000.00
    # Total sum formula row (row 3)
    assert ws_invoices["D3"].value == "Total:"
    assert ws_invoices["E3"].value == "=SUM(E2:E2)"

    # Check summary sheet formulas
    ws_summary = wb["Summary"]
    assert ws_summary["A1"].value == "Invoice Processing Report"
    assert ws_summary["B5"].value == "=COUNTA(Invoices!A2:A2)"
    assert ws_summary["B6"].value == "=SUM(Invoices!E2:E2)"
