"""
Excel utilities wrapping pandas and openpyxl with dynamic imports and error handling.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any

from ..exceptions import ExcelError

if TYPE_CHECKING:
    import pandas as pd


class ExcelReader:
    """Reads Excel files into pandas DataFrames."""

    @staticmethod
    def read_sheet_to_df(
        path: str | Path,
        sheet_name: str | int = 0,
        **kwargs: Any,
    ) -> pd.DataFrame:
        """
        Read an Excel sheet into a pandas DataFrame.
        Import pandas dynamically to keep dependencies optional.
        """
        try:
            import pandas as pd
        except ImportError as exc:
            raise ExcelError(
                "pandas is required for reading Excel files but not installed. "
                "Install it using `pip install pandas openpyxl` or use extra `data`.",
            ) from exc

        filepath = Path(path)
        if not filepath.exists():
            raise ExcelError(f"Excel file not found at path: {filepath}")

        try:
            # Use openpyxl as default engine for modern .xlsx files
            return pd.read_excel(filepath, sheet_name=sheet_name, engine="openpyxl", **kwargs)
        except Exception as exc:
            raise ExcelError(f"Failed to read Excel file {filepath.name}: {exc}") from exc


class ExcelWriter:
    """Writes DataFrames to Excel files with styling."""

    @staticmethod
    def write_df_to_sheet(
        df: pd.DataFrame,
        path: str | Path,
        sheet_name: str = "Sheet1",
        auto_fit: bool = True,
        style_header: bool = True,
    ) -> None:
        """
        Write a pandas DataFrame to an Excel sheet.
        Optionally format headers and auto-fit column widths.
        """
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError as exc:
            raise ExcelError(
                "pandas and openpyxl are required for writing Excel files but not installed. "
                "Install them using `pip install pandas openpyxl` or use extra `data`.",
            ) from exc

        filepath = Path(path)
        filepath.parent.mkdir(parents=True, exist_ok=True)

        try:
            # Create or load workbook
            if filepath.exists():
                wb = load_workbook(filepath)
                # Remove sheet if it already exists to overwrite
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                ws = wb.create_sheet(sheet_name)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name

            # Make sure gridlines are visible
            ws.views.sheetView[0].showGridLines = True

            # Write DataFrame to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    # Styling header
                    if r_idx == 1 and style_header:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # Apply a thin default border to all cells
            thin_border = Border(
                left=Side(style="thin", color="D3D3D3"),
                right=Side(style="thin", color="D3D3D3"),
                top=Side(style="thin", color="D3D3D3"),
                bottom=Side(style="thin", color="D3D3D3"),
            )
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            # Auto-fit column widths
            if auto_fit:
                for col in ws.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            wb.save(filepath)
        except Exception as exc:
            raise ExcelError(f"Failed to write Excel file {filepath.name}: {exc}") from exc
